from __future__ import print_function
import requests
import sys
import ssl
import urllib.request as req
from urllib.request import urlretrieve
from bs4 import BeautifulSoup
from tempfile import NamedTemporaryFile
import openpyxl as excel
import os
from datetime import datetime, date, time
import re
import pickle
import os.path
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request

# If modifying these scopes, delete the file token.pickle.
SCOPES = ['https://www.googleapis.com/auth/spreadsheets']

TRADES_SPREADSHEET_ID = '1bevgPBYdh6-hHqFGKQ6o7cBLsFaY15yHL9PLQPtd3ks'

CARS_URL = 'https://xn----etbpba5admdlad.xn--p1ai/bankrot?categorie_childs%5B0%5D=2&regions%5B0%5D=50&regions%5B1%5D=77&section=%D0%91%D0%B0%D0%BD%D0%BA%D1%80%D0%BE%D1%82%D1%81%D1%82%D0%B2%D0%BE&forms%5B0%5D=public&forms%5B1%5D=auction&page='

CARS_EN_URL = 'https://xn----etbpba5admdlad.xn--p1ai/bankrot?categorie_childs%5B0%5D=2&regions%5B0%5D=50&regions%5B1%5D=77&section=%D0%91%D0%B0%D0%BD%D0%BA%D1%80%D0%BE%D1%82%D1%81%D1%82%D0%B2%D0%BE&forms%5B0%5D=public&forms%5B1%5D=auction&page=1'

vinregex = '[0-9abcdefghjklmnprstuvwxyzABCDEFGHJKLMNPRSTUVWXYZ]{17,20}'


 #-*- coding: utf-8 -*-


def google_auth():
    """Shows basic usage of the Sheets API.
    Prints values from a sample spreadsheet.
    """
    creds = None
    # The file token.pickle stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    if os.path.exists('token.pickle'):
        with open('token.pickle', 'rb') as token:
            creds = pickle.load(token)
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open('token.pickle', 'wb') as token:
            pickle.dump(creds, token)

    service = build('sheets', 'v4', credentials=creds)
    return service

def get_sheet(service, sheet, srange):
    # Call the Sheets API

    resrange = sheet+'!'+srange

    sheet = service.spreadsheets()
    result = sheet.values().get(spreadsheetId=TRADES_SPREADSHEET_ID,
                                range=resrange).execute()
    values = result.get('values', [])
    return values
 #   if not values:
 #       print('No data found.')
 #   else:
 #       print('Name, Major:')
 #       for row in values:
 #           # Print columns A and E, which correspond to indices 0 and 4.
 #           print('%s, %s' % (row[1], row[2]))

def clear(sheet, service):
    clear = [' ',' ',' ',' ',' ',' ',' ',' ',' ',' ']
    body = {
    'range' : sheet+'!'+'A2:J',
    }
    result = service.spreadsheets().values().clear(spreadsheetId=TRADES_SPREADSHEET_ID, range=sheet+'!'+'A2:J', body=body)
    result.execute()

def write_sheet(service, sheet, srange, data):
    resrange = sheet+'!'+srange
    sheet = service.spreadsheets()
    values = []
    for car in data:
        value = []
        value.append(car['id'])
        value.append(car['name'])
        value.append(car['act_price'])
        value.append(car['start_price'])
        value.append(car['start'])
        value.append(car['end'])
        value.append(car['link'])
        value.append(car['type'])
        value.append(car['vins'])
        value.append(car['info'])
        values.append(value)
    body = {
    'values': values,
    'range' : resrange,
    'majorDimension':'ROWS'
    }
    result = service.spreadsheets().values().update(spreadsheetId=TRADES_SPREADSHEET_ID, range=resrange, valueInputOption='RAW', body=body)
    result.execute()

def fetch_torgi_page():
    page = NamedTemporaryFile()
    urlretrieve(CARS_EN_URL, page.name)
    return page

def fetch_one_page(pgnum):
    page = NamedTemporaryFile()
    urlretrieve(CARS_URL+str(pgnum), page.name)
    return page

def fetch_url(addr):
    page = NamedTemporaryFile()
    urlretrieve(addr, page.name)
    return page


def get_car_info_from_div(div):
    block_divs = div.find_all('div')
    car_id = get_car_id(block_divs)
    car_name = get_car_name(block_divs)
    car_info = get_car_info(block_divs)
    car_link = get_car_link(block_divs)
    car = fetch_url(car_link)
    car_divs, car_p = parse_car_page(car)
    car_act_price, car_start_price = get_car_price(car_divs)
    auction_type = get_car_auction_type(block_divs)
    d_start = get_date_start(car_p)
    d_end = get_date_end(car_p)
    vins = get_vin(car_info, car_name)
    #d_start =
    #d_end = 
    #trade_place = 
    return {'id': car_id, 'name': car_name, 'act_price': car_act_price, 'start_price': car_start_price, 'start': d_start, 'end':d_end, 'link': car_link, 'type': auction_type, 'vins':vins, 'info': car_info}

def get_vin(car_info, car_name):
    text3 = car_info+' '+car_name
    vins = []
    replacer = ['А', 'A', 'В', 'B', 'Е', 'E', 'К', 'K', 'М', 'M','Н', 'H','О', 'O','Р', 'P','С', 'C','Т', 'T', 'У', 'Y', 'Х', 'X']
    ru = []
    en = []
    text3.upper()
    i = 0
    while i < len(replacer)-1:
        ru.append(replacer[i])
        en.append(replacer[i+1])
        i = i+2
    i = 0
    while i < len(text3):
        if text3[i] in ru:
            text3 = text3[0:i]+str(en[ru.index(text3[i])])+text3[i+1:]
        i=i+1
    res = re.findall(vinregex, text3)
    for vin in res:
        if vin not in vins:
            vins.append(vin)
    return str(vins)

def get_cars_from_torgi(divs):
    car_divs = list(filter(lambda div: 'class' in div.attrs and
                                         'lot-card-wrapper' in div.get('class'), divs))
    cars = list(map(lambda div: get_car_info_from_div(div), car_divs))
    return cars

def get_date_start(p_tags):
            if p_tags[4].text[0:6] == 'Начало':
                return p_tags[4].text[21:31]
            if p_tags[5].text[0:6] == 'Начало':
                return p_tags[5].text[21:31]

def get_date_end(p_tags):
            if p_tags[5].text[0:5] == 'Конец':
                return p_tags[5].text[20:30]
            if p_tags[6].text[0:5] == 'Конец':
                return p_tags[6].text[20:30]



def get_car_id(div_tags):
    for car_div in div_tags:
        if 'class' in car_div.attrs and \
                        'component4__header' in car_div.get('class'):
            return car_div.span.string

def get_car_link(div_tags):
    for car_div in div_tags:
        if 'class' in car_div.attrs and \
                        'component4__body' in car_div.get('class'):
            return car_div.a.get('href')

def get_car_name(div_tags):
    for car_div in div_tags:
        if 'class' in car_div.attrs and \
                        'component4__body' in car_div.get('class'):
            return car_div.h3.string

def get_car_info(div_tags):
    for car_div in div_tags:
        if 'class' in car_div.attrs and \
                        'component4__body' in car_div.get('class'):
                try:        
                    return car_div.p.string
                except AttributeError:
                    return ""
                     

def get_car_price(div_tags):
    for price_div in div_tags:
        if 'class' in price_div.attrs and \
                        'new-component1__price' in price_div.get('class'):
            price_list = list(price_div)
            if len(price_list)>3:
                return price_list[1].text, price_list[3].text
            else:
                return price_list[1].text, price_list[1].text,

def get_car_auction_type(div_tags):
    for car_div in div_tags:
        if 'class' in car_div.attrs and \
                        'new-component1' in car_div.get('class'):
            try:
                return car_div.a.img.get('data-tooltip')
            except AttributeError:
                return ""

def parse_cars_list(raw_html_file):
    soup = BeautifulSoup(raw_html_file.read(), 'html.parser')
    div_tags = soup.find_all('div')
    return div_tags

def parse_pages_list(raw_html_file):
    soup = BeautifulSoup(raw_html_file.read(), 'html.parser')
    li_tags = soup.find_all('li')
    return li_tags

def parse_page(raw_html_file):
    soup = BeautifulSoup(raw_html_file.read(), 'html.parser')
    div_tags = soup.find_all('div')
    li_tags = soup.find_all('li')
    return div_tags, li_tags

def parse_car_page(raw_html_file):
    soup = BeautifulSoup(raw_html_file.read(), 'html.parser')
    div_tags = soup.find_all('div')
    p_tags = soup.find_all('p')
    return div_tags, p_tags


def get_page_count(divs):
    pages_list = list(filter(lambda li: 'class' in li.attrs and 'page-item' in li.get('class'), divs))
   # pagination = list(filter(lambda div: 'class' in div.attrs and 'pagination' in div.get('class'), divs))
   # pages_count = len(pagination.ul.find_all('li'))
   # return pages_count
    return len(pages_list)-2

def get_cars_on_page(page):
    page_name = fetch_one_page(page)
    page_div_tags, page_li_tags = parse_page(page_name)
    cars_on_page = get_cars_from_torgi(page_div_tags)
    return cars_on_page

def put_to_excel(cars):
    filename = os.path.abspath('Trades.xlsx')
    workbook = excel.load_workbook(filename)
    ws = workbook['LastDownload']
    target = workbook.copy_worksheet(ws)
    target.title = 'Download'+str(datetime.date(datetime.today()))
    for row in ws['A2:J'+str(ws.max_row)]:
        for cell in row:
            cell.value = None
    row = 2 
    for car in cars:
        ws.cell(row=row, column=1, value=car['id'])
        ws.cell(row=row, column=2, value=car['name'])
        ws.cell(row=row, column=3, value=car['act_price'])
        ws.cell(row=row, column=4, value=car['start_price'])
        ws.cell(row=row, column=5, value=car['start'])
        ws.cell(row=row, column=6, value=car['end'])
        ws.cell(row=row, column=7, value=car['link'])
        ws.cell(row=row, column=8, value=car['type'])
        ws.cell(row=row, column=9, value=car['vins'])
        ws.cell(row=row, column=10, value=car['info'])
        row = row +1
    workbook.save(filename)
    workbook.close()


def put_new_to_excel(cars):
    filename = os.path.abspath('Trades.xlsx')
    workbook = excel.load_workbook(filename)
    ws = workbook['New']
    for row in ws['A2:J'+str(ws.max_row)]:
        for cell in row:
            cell.value = None
    row = 2 
    for car in cars:
        ws.cell(row=row, column=1, value=car['id'])
        ws.cell(row=row, column=2, value=car['name'])
        ws.cell(row=row, column=3, value=car['act_price'])
        ws.cell(row=row, column=4, value=car['start_price'])
        ws.cell(row=row, column=5, value=car['start'])
        ws.cell(row=row, column=6, value=car['end'])
        ws.cell(row=row, column=7, value=car['link'])
        ws.cell(row=row, column=8, value=car['type'])
        ws.cell(row=row, column=9, value=car['vins'])
        ws.cell(row=row, column=10, value=car['info'])
        row = row +1
    workbook.save(filename)
    workbook.close()



def read_existing_lots(service):
    rows = get_sheet(service, 'LastDownload', 'A2:A')
    id = []
    for row in rows:
        i = 0
        while i < len(row):
            id.append(row[i])
            i = i+1
    return id

def new_cars(cars, ids):
    newcars =[]
    for car in cars:
        if str(car['id']) in ids:
            pass
        else:
            newcars.append(car)
    return newcars



if __name__ == '__main__':
    ssl._create_default_https_context = ssl._create_unverified_context
    html_page_name = fetch_torgi_page()
    page_div_tags, page_li_tags = parse_page(html_page_name)
    pages_count = get_page_count(page_li_tags)
    serv = google_auth()
    clear('New', serv)
    print('Pages:'+str(pages_count))
    i = 1
    cars=[]
    while i <= pages_count:
        print('Page: '+str(i))
        cars.extend(get_cars_on_page(i))
        i = i+1
    ids = read_existing_lots(serv)
    mycars = new_cars(cars, ids)
    clear('New', serv)
    write_sheet(serv, 'New', 'A2:J'+str(len(mycars)+1), mycars)
    clear('LastDownload',serv)
    write_sheet(serv, 'LastDownload', 'A2:J'+str(len(cars)+1), cars)
    #put_new_to_excel(mycars)
    #put_to_excel(cars)
    print('Finished!')